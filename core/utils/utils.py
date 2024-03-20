import openpyxl
import random


book = openpyxl.load_workbook(filename='schedule.xlsx')
ws = book.active


def get_column_for_class(num_and_letter: str) -> int:
    try:
        for i in range(1, ws.max_row):
            current_cell = ws.cell(row=4, column=i)
            if str(current_cell.value).lower() == num_and_letter:
                index_current_cell = current_cell.column
                # print('–ù–∞–π–¥–µ–Ω–Ω—ã–π —Å—Ç–æ–ª–±–∏–∫: ', current_cell.column, '\n–ê–¥—Ä–µ—Å —è—á–µ–π–∫–∏: ', current_cell.coordinate)
                return index_current_cell

        print('–¢–∞–∫–æ–≥–æ –∫–ª–∞—Å—Å–∞ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç')
    except:
        print('—á—Ç–æ-—Ç–æ –Ω–µ —Ç–∞–∫ —Å –ø–æ–∏—Å–∫–æ–º –Ω—É–∂–Ω–æ–≥–æ –∫–ª–∞—Å—Å–∞')


def get_day_of_week_row(day_of_week: str) -> int:
    try:
        for i in range(1, ws.max_column):
            current_cell = ws.cell(row=i, column=1)
            if str(current_cell.value).lower() == day_of_week:
                # print('–ù–∞–π–¥–µ–Ω–Ω–∞—è —Å—Ç—Ä–æ–∫–∞: ', current_cell.row)
                return current_cell.row
    except:
        print('–î–µ–Ω—å –Ω–µ —É–¥–∞—ë—Ç—Å—è –Ω–∞–π—Ç–∏')


def get_schedule(column, row):
    lessons_list = []
    lessons_room = []
    try:
        for i in range(7):
            cur_cell = ws.cell(row=row + i, column=column)
            cur_cell_room = ws.cell(row=row + i, column=column+1)
            if cur_cell.value is None:
                continue
            if cur_cell_room.value is None:
                continue

            lessons_list.append(cur_cell.value)
            lessons_room.append(cur_cell_room.value)

        return lessons_list, lessons_room
    except:
        return None


# print(get_schedule(get_column_for_class('5–±'), get_day_of_week_row('–ø–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫')))


def parser_of_data_hw(data):
    list_of_emoji = ['‚ö†', 'üìé', 'üìå', 'üî¨üß¨', 'üóø', 'üìµ', 'üÜò']
    result_string = ''
    for homework in data:
        for target in homework:
            cur_emoji = random.choice(list_of_emoji)
            #–∑–¥–µ—Å—å –º–æ–∂–Ω–æ –æ—Ç–ø—Ä–∞–≤–ª—è—Ç—å –¥–∞—Ç—É —Å–æ–∑–¥–∞–Ω–∏—è –¥–∑
            result_string += f'{cur_emoji*5}\n'
            result_string += f'{str(target)}\n'
    return result_string


# –Ω–∞–ø–∏—Å–∞—Ç—å —Ñ—É–∫–Ω—Ü–∏—é –¥–ª—è –≤—ã–≤–æ–¥–∞ —Ü–µ–ª–æ–π –Ω–µ–¥–µ–ª–∏ –Ω–∞ –∫–ª–∞—Å—Å
# –Ω—É–∂–Ω–æ –Ω–∞–ø–∏—Å–∞—Ç—å —Ñ—É–∫—Ü–∏—é, –∫–æ—Ç–æ—Ä–∞—è –±—É–¥–µ—Ç –∫–∏–¥–∞—Ç—å —Ä–∞—Å—Å–ø–∏—Å–∞–Ω–∏–µ –ø–æ —Ç–∞–π–º–µ—Ä—É –Ω–∞ –∫–∞–∂–¥—ã–π –¥–µ–Ω—å
# –Ω–∞–ø–∏—Å–∞—Ç—å —Ñ—É–Ω—Ü–∏–∏ –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–∑ –∏ –µ–≥–æ –∏–∑–º–µ–Ω–µ–Ω–∏—è


# def get_schedule_for_all_week(num_and_letter: str):
#     d1 = get_schedule(get_column_for_class(num_and_letter), get_day_of_week_row('–ø–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫'))
#     d2 = get_schedule(get_column_for_class(num_and_letter), get_day_of_week_row('–≤—Ç–æ—Ä–Ω–∏–∫'))
#     d3 = get_schedule(get_column_for_class(num_and_letter), get_day_of_week_row('—Å—Ä–µ–¥–∞'))
#     d4 = get_schedule(get_column_for_class(num_and_letter), get_day_of_week_row('—á–µ—Ç–≤–µ—Ä–≥'))
#     d5 = get_schedule(get_column_for_class(num_and_letter), get_day_of_week_row('–ø—è—Ç–Ω–∏—Ü–∞'))
#     d6 = get_schedule(get_column_for_class(num_and_letter), get_day_of_week_row('—Å—É–±–±–æ—Ç–∞'))
#
#     print(
#         '–ø–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫ \n', d1,
#           '\n–≤—Ç–æ—Ä–Ω–∏–∫ \n', d2,
#           '\n—Å—Ä–µ–¥–∞ \n', d3,
#           '\n—á–µ—Ç–≤–µ—Ä–≥ \n', d4,
#           '\n–ø—è—Ç–Ω–∏—Ü–∞ \n', d5,
#           '\n—Å—É–±–±–æ—Ç–∞ \n', d6)
#
#
# get_schedule_for_all_week('5–±')