import openpyxl


book = openpyxl.load_workbook(filename='schedule.xlsx')
ws = book.active

current_cell = ws.cell(row=4, column=6)

my_list = [45,5,4,32,5,2,35,231,4,12321,12,3123, 'ddfsfs' ,4,5235 ,7 , 21332,423423,45, 7, 7, '7']
for i in range(len(my_list)):
    try:
        if int(my_list[i]) == 7:
            print(my_list[i])
    except:
        print("Скрипт не смог конвертировать елемент ", my_list[i])


def get_column_for_class(num_and_letter: str) -> int:
    for i in range(1, ws.max_row):
        current_cell = 0


# print(current_cell)